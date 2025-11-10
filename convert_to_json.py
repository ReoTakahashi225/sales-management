# -*- coding: utf-8 -*-
import openpyxl
import json
from datetime import datetime

# Excelファイルを開く
wb = openpyxl.load_workbook('【master】年間目標_21期_実績_202510.xlsx', data_only=True)

print("利用可能なシート:")
for i, sheet_name in enumerate(wb.sheetnames):
    print(f"  {i}: {sheet_name}")

# 社員売上シートを読み込む (インデックスで指定)
ws = wb.worksheets[2]  # 3番目のシート「社員売上」

# 月のキー定義
MONTH_KEYS = ['oct', 'nov', 'dec', 'jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep']

# データ構造
employees = []
partners = []
expenses = {}
targets = {}

# 月ごとの支出データを初期化
for key in MONTH_KEYS:
    expenses[key] = {
        'salary': 0,
        'insurance': 0,
        'personalExpense': 0
    }
    targets[key] = 0  # 目標売上の初期値

print(f"\n{ws.title}シートを読み込み中...")

# 行データを取得 (5行目からスタート - 氏名データ)
row_data = list(ws.iter_rows(min_row=5, values_only=True))

for idx, row in enumerate(row_data, start=5):
    # 空行をスキップ
    if not row[1] or row[1] is None:
        continue

    # 社員名を取得 (2列目、インデックス1)
    name = str(row[1]).strip() if row[1] else None

    if not name or name == 'None':
        continue

    print(f"処理中: 行{idx}, 氏名={name}")

    # 各月のデータを取得
    # 列の構造: No, 氏名, 10月(顧客,売上,BSG,形態), 11月(...), ...
    # 顧客: col 2, 売上: col 3, 形態: col 5
    # 次の月: 顧客: col 6, 売上: col 7, 形態: col 9

    employee_data = {
        'id': str(int(datetime.now().timestamp() * 1000) + idx),
        'name': name,
        'dept': '',  # 所属はExcelに含まれていないため空
        'updatedAt': datetime.now().isoformat()
    }

    # 各月のデータ
    for month_idx, month_key in enumerate(MONTH_KEYS):
        # 列のインデックスを計算
        # 10月は列2から始まる (顧客=2, 売上=3, BSG=4, 形態=5)
        base_col = 2 + (month_idx * 4)

        customer = str(row[base_col]).strip() if base_col < len(row) and row[base_col] else ''
        revenue = row[base_col + 1] if base_col + 1 < len(row) and row[base_col + 1] else 0
        bsg = str(row[base_col + 2]).strip() if base_col + 2 < len(row) and row[base_col + 2] else ''
        work_type = str(row[base_col + 3]).strip() if base_col + 3 < len(row) and row[base_col + 3] else ''

        # 売上が数値でない場合は0にする
        if not isinstance(revenue, (int, float)):
            revenue = 0

        employee_data[month_key] = {
            'customer': customer if customer and customer != 'None' else '',
            'revenue': int(revenue) if revenue else 0,
            'type': work_type if work_type and work_type != 'None' else ''
        }

    # 年間売上が0でない場合のみ追加
    total_revenue = sum(employee_data[key]['revenue'] for key in MONTH_KEYS)
    if total_revenue > 0:
        # BSGやSYSGなどを部署として設定
        for key in MONTH_KEYS:
            if employee_data[key]['customer']:
                # 最初に見つかった顧客情報を部署にする
                dept_info = ''
                if base_col + 2 < len(row) and row[base_col + 2]:
                    dept_info = str(row[base_col + 2]).strip()
                if dept_info and dept_info != 'None':
                    employee_data['dept'] = dept_info
                break

        employees.append(employee_data)
        print(f"  追加: {name}, 年間売上={total_revenue:,}円")

print(f"\n社員データ: {len(employees)}件")

# BPシートを読み込む (インデックス4)
ws_bp = wb.worksheets[4]  # 5番目のシート「BP原価管理」
print(f"\n{ws_bp.title}シートを読み込み中...")

# 最初の10行を確認
print("\n最初の10行を確認:")
test_rows = list(ws_bp.iter_rows(min_row=1, max_row=10, values_only=True))
for i, row in enumerate(test_rows, 1):
    print(f"  行{i}: {row[:5] if len(row) > 5 else row}")

row_data_bp = list(ws_bp.iter_rows(min_row=3, values_only=True))

for idx, row in enumerate(row_data_bp, start=3):
    # 空行チェック
    if not row or len(row) < 3:
        continue

    # BP名を取得 (列インデックス2 - 3列目)
    name_val = row[2]
    company_val = ''  # 所属会社情報がない場合があるので空で初期化

    # 数値や不要なデータをスキップ
    if name_val is None or name_val == '' or isinstance(name_val, (int, float)):
        continue

    name = str(name_val).strip()

    # ヘッダー行や不要なデータをスキップ
    if name in ['BP名管理', 'None', '所属', '所属会社']:
        continue

    print(f"処理中: 行{idx}, BP名={name}")

    partner_data = {
        'id': str(int(datetime.now().timestamp() * 1000) + idx + 10000),
        'name': name,
        'company': company_val,
        'updatedAt': datetime.now().isoformat()
    }

    # 各月のデータ
    # 列構造: No, (空?), BP名, 10月(売上,請求額,粗利), 11月(...), ...
    # 請求額は列4 (index 4), 7, 10, ... (3列おき)
    for month_idx, month_key in enumerate(MONTH_KEYS):
        # 列インデックス: 4, 7, 10, 13, ...
        cost_col = 4 + (month_idx * 3)
        cost = row[cost_col] if cost_col < len(row) and row[cost_col] else 0

        if not isinstance(cost, (int, float)):
            cost = 0

        partner_data[month_key] = {
            'cost': int(abs(cost)) if cost else 0  # 絶対値を取る (負の値の場合)
        }

    # 年間コストが0でない場合のみ追加
    total_cost = sum(partner_data[key]['cost'] for key in MONTH_KEYS)
    if total_cost > 0:
        partners.append(partner_data)
        print(f"  追加: {name}, 年間コスト={total_cost:,}円")
    else:
        print(f"  スキップ: {name}, 年間コスト=0円")

print(f"\nBPデータ: {len(partners)}件")

# JSONデータを作成
output_data = {
    'employees': employees,
    'partners': partners,
    'expenses': expenses,
    'targets': targets,
    'exportedAt': datetime.now().isoformat(),
    'version': '2.1',
    'period': '第21期（2025年10月～2026年9月）'
}

# JSONファイルに保存
output_file = '売上実績データ.json'
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(output_data, f, ensure_ascii=False, indent=2)

print(f"\nJSONファイルを作成しました: {output_file}")
print(f"  社員: {len(employees)}件")
print(f"  BP: {len(partners)}件")

# プレビューを表示
print("\n--- データプレビュー ---")
if employees:
    print(f"\n社員例1: {employees[0]['name']}")
    print(f"  部署: {employees[0]['dept']}")
    print(f"  10月: 顧客={employees[0]['oct']['customer']}, 売上={employees[0]['oct']['revenue']:,}円, 形態={employees[0]['oct']['type']}")
    if len(employees) > 1:
        print(f"\n社員例2: {employees[1]['name']}")
        print(f"  部署: {employees[1]['dept']}")
        print(f"  10月: 顧客={employees[1]['oct']['customer']}, 売上={employees[1]['oct']['revenue']:,}円, 形態={employees[1]['oct']['type']}")

if partners:
    print(f"\nBP例: {partners[0]['name']}")
    print(f"  所属: {partners[0]['company']}")
    print(f"  10月: 請求={partners[0]['oct']['cost']:,}円")
